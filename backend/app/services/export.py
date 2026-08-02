from collections.abc import Sequence
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models import RegistryObject
from app.parsers.registry import REGISTRY_HEADERS, registry_export_row


EMPLOYEES = [
    "Джерелиевский Д.Б.",
    "Хлынцева Л.А",
    "Непомнящая Л.В.",
    "Смоляницкая А.И.",
    "Смоляницкий А.Г.",
    "Гуров М.В.",
    "Попова М.М.",
    "Сотников С.А.",
    "Нечитайлова Н.А.",
    "Брисева А.С.",
    "Субботин В.М.",
]

DNA_METHODS = [
    'Maxwell 16 "LEV"',
    "QIAamp DNA Investigator Kit",
    "QIAamp DNA FFPE Tissue Kit",
    "PrepFiler Express Forensic DNA",
    "Prep-n-Go",
    'АТГ "Биотех"',
    "QIACUBE",
    "Реагент SwabSolution Kit 100rxn",
    "Набор для выделения CordiS Экстракт",
    "Casework Direct Kit",
    "QIAsymphony DNA InvestigatorKit",
    "PrepFiler Manual DNA Extraction Kit",
    "Chelex",
    "RapidHiT",
    "MagaBio DNA Purification KIT_32",
    "GeneRead DNA FFPE",
    "М-Сорб кость",
]


def build_registry_workbook(objects: Sequence[RegistryObject]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "список объектов"
    ws.append(REGISTRY_HEADERS)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for obj in objects:
        ws.append(registry_export_row(obj))
    ws.auto_filter.ref = f"A1:{get_column_letter(len(REGISTRY_HEADERS))}{max(ws.max_row, 1)}"
    ws.freeze_panes = "A2"
    for col_idx in range(1, len(REGISTRY_HEADERS) + 1):
        width = 14
        if col_idx in {8, 9}:
            width = 42
        elif col_idx > 17:
            width = 24
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if hasattr(cell.value, "strftime"):
                cell.number_format = "DD.MM.YYYY"

    data = wb.create_sheet("Данные для вып.окна")
    data.append(["Столбец1", "", "Столбец1"])
    max_len = max(len(EMPLOYEES), len(DNA_METHODS))
    for idx in range(max_len):
        data.append([
            EMPLOYEES[idx] if idx < len(EMPLOYEES) else "",
            "",
            DNA_METHODS[idx] if idx < len(DNA_METHODS) else "",
        ])
    data.column_dimensions["A"].width = 32
    data.column_dimensions["C"].width = 42

    empty = wb.create_sheet("Пустой")
    empty.append(["Лабораторный номер", "Основа", "№ рег РЦСМЭ", "№ постановления"])

    output = BytesIO()
    wb.save(output)
    return output.getvalue()
