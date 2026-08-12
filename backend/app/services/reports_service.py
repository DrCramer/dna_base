from __future__ import annotations

from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from io import BytesIO
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import Integer, case, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import ElectrophoresisControlFile, ElectrophoresisResultFile, Party, RegistryObject, StageEvent, StageEventPerformer
from app.services.no_object import control_no_tokens, control_token_matches_object, has_no_object_marker


LAB_STAGES = ["sample_prep", "milling", "dna_extraction", "realtime", "pcr", "electrophoresis", "analysis"]
REPORT_STAGE_LABELS = {
    "registration": "Регистрация",
    "sample_prep": "Пробоподготовка",
    "milling": "Измельчение",
    "dna_extraction": "Выделение",
    "realtime": "RealTime",
    "pcr": "ПЦР",
    "electrophoresis": "Электрофорез",
    "analysis": "Анализ",
}
CONTROL_FIELDS = [
    "control_actual_decrees",
    "control_decree_without_object",
    "control_object_without_decree",
    "control_unidentified_rostov_no",
    "control_need_recall",
    "control_recalled",
]
CONTROL_LABELS = {
    "control_actual_decrees": "Фактическое количество постановлений",
    "control_decree_without_object": "Есть постановление, но нет объекта",
    "control_object_without_decree": "Есть объект, но нет постановления",
    "control_unidentified_rostov_no": "Неидентифицируемый ростовский номер",
    "control_need_recall": "Надо отозвать",
    "control_recalled": "Отозваны",
}


@dataclass(slots=True)
class ReportFilters:
    case_year: int | None = None
    date_from: date | None = None
    date_to: date | None = None
    party_ids: list[int] | None = None
    stage_type: str | None = None
    employee_id: int | None = None
    object_type: str | None = None
    box_no: str | None = None
    include_archived: bool = False
    include_empty_parties: bool = True
    only_problematic: bool = False
    page: int = 1
    page_size: int = 200
    sort_by: str | None = None
    sort_dir: str = "desc"


def _filled(value: Any) -> bool:
    return bool(str(value or "").strip())


def _token_count(value: str | None) -> int:
    return len(control_no_tokens(value))


def _control_status(party: Party) -> str:
    has_any = any(_filled(getattr(party, field)) for field in CONTROL_FIELDS)
    if not has_any:
        return "Нет данных"
    if _filled(party.control_need_recall) or _filled(party.control_recalled):
        return "Критично"
    if (
        _filled(party.control_decree_without_object)
        or _filled(party.control_object_without_decree)
        or _filled(party.control_unidentified_rostov_no)
    ):
        return "Есть замечания"
    return "Без проблем"


def _control_problem_count(party: Party) -> int:
    return (
        _token_count(party.control_decree_without_object)
        + _token_count(party.control_object_without_decree)
        + _token_count(party.control_unidentified_rostov_no)
        + _token_count(party.control_need_recall)
        + _token_count(party.control_recalled)
    )


def _is_burnt_bone(description: str | None) -> bool:
    text = (description or "").casefold()
    return "горел" in text and "кость" in text


def _is_no_biomaterial(description: str | None) -> bool:
    text = (description or "").casefold()
    return "нет биоматериала" in text


def _numeric_party_no_expr():
    return case((Party.party_no.op("~")(r"^\d+$"), Party.party_no.cast(Integer)), else_=None)


def _page_rows(rows: list[dict[str, Any]], filters: ReportFilters) -> dict[str, Any]:
    total = len(rows)
    page_size = max(1, min(filters.page_size or 200, 1000))
    page = max(1, filters.page or 1)
    start = (page - 1) * page_size
    return {"items": rows[start:start + page_size], "total": total, "page": page, "page_size": page_size}


def _sort_rows(rows: list[dict[str, Any]], filters: ReportFilters, default: str = "party_no_numeric") -> list[dict[str, Any]]:
    key = filters.sort_by or default
    reverse = (filters.sort_dir or "desc").lower() != "asc"

    def value(row: dict[str, Any]):
        raw = row.get(key)
        if raw is None and key == "party_no_numeric":
            raw = row.get("party_no_sort")
        if isinstance(raw, str):
            return raw.casefold()
        return raw if raw is not None else -1

    return sorted(rows, key=value, reverse=reverse)


def _percent(done: int, total: int) -> int:
    if total <= 0:
        return 0
    return round((done / total) * 100)


async def _filtered_parties(session: AsyncSession, filters: ReportFilters) -> list[Party]:
    stmt = select(Party)
    if not filters.include_archived:
        stmt = stmt.where(Party.status != "archived")
    if filters.case_year:
        stmt = stmt.where(Party.case_year == filters.case_year)
    if filters.party_ids:
        stmt = stmt.where(Party.id.in_(filters.party_ids))
    if filters.object_type or filters.box_no:
        object_party_ids = select(RegistryObject.party_id).where(RegistryObject.status != "archived")
        if filters.object_type:
            object_party_ids = object_party_ids.where(RegistryObject.object_type.ilike(f"%{filters.object_type.strip()}%"))
        if filters.box_no:
            object_party_ids = object_party_ids.where(RegistryObject.box_no.ilike(f"%{filters.box_no.strip()}%"))
        stmt = stmt.where(Party.id.in_(object_party_ids))
    numeric = _numeric_party_no_expr()
    stmt = stmt.order_by(numeric.desc().nullslast(), Party.party_no.desc(), Party.id.desc())
    result = await session.execute(stmt)
    return list(result.scalars().all())


def _object_filter_conditions(filters: ReportFilters) -> list[Any]:
    conditions: list[Any] = []
    if filters.object_type:
        conditions.append(RegistryObject.object_type.ilike(f"%{filters.object_type.strip()}%"))
    if filters.box_no:
        conditions.append(RegistryObject.box_no.ilike(f"%{filters.box_no.strip()}%"))
    return conditions


async def _party_object_counts(
    session: AsyncSession,
    party_ids: list[int],
    filters: ReportFilters,
    include_archived_objects: bool = False,
) -> dict[int, int]:
    if not party_ids:
        return {}
    stmt = select(RegistryObject.party_id, func.count(RegistryObject.id)).where(RegistryObject.party_id.in_(party_ids))
    if not include_archived_objects:
        stmt = stmt.where(RegistryObject.status != "archived")
    for condition in _object_filter_conditions(filters):
        stmt = stmt.where(condition)
    stmt = stmt.group_by(RegistryObject.party_id)
    rows = await session.execute(stmt)
    return {int(party_id): int(count) for party_id, count in rows.all() if party_id is not None}


async def _party_stage_counts(session: AsyncSession, party_ids: list[int], filters: ReportFilters) -> tuple[dict[int, dict[str, int]], dict[int, datetime | None], dict[int, int]]:
    if not party_ids:
        return {}, {}, {}
    stage_stmt = (
        select(RegistryObject.party_id, StageEvent.stage_type, func.count(func.distinct(StageEvent.object_id)))
        .join(StageEvent, StageEvent.object_id == RegistryObject.id)
        .where(
            RegistryObject.party_id.in_(party_ids),
            RegistryObject.status != "archived",
            StageEvent.is_cancelled.is_(False),
        )
        .group_by(RegistryObject.party_id, StageEvent.stage_type)
    )
    for condition in _object_filter_conditions(filters):
        stage_stmt = stage_stmt.where(condition)
    rows = await session.execute(stage_stmt)
    counts: dict[int, dict[str, int]] = defaultdict(dict)
    for party_id, stage_type, count in rows.all():
        if party_id is not None:
            counts[int(party_id)][stage_type] = int(count)
    latest_stmt = (
        select(RegistryObject.party_id, func.max(StageEvent.updated_at))
        .join(StageEvent, StageEvent.object_id == RegistryObject.id)
        .where(
            RegistryObject.party_id.in_(party_ids),
            RegistryObject.status != "archived",
            StageEvent.is_cancelled.is_(False),
        )
        .group_by(RegistryObject.party_id)
    )
    for condition in _object_filter_conditions(filters):
        latest_stmt = latest_stmt.where(condition)
    latest_rows = await session.execute(latest_stmt)
    latest = {int(party_id): updated_at for party_id, updated_at in latest_rows.all() if party_id is not None}
    repeat_stmt = (
        select(RegistryObject.party_id, func.count(func.distinct(StageEvent.object_id)))
        .join(StageEvent, StageEvent.object_id == RegistryObject.id)
        .where(
            RegistryObject.party_id.in_(party_ids),
            RegistryObject.status != "archived",
            StageEvent.is_cancelled.is_(False),
            StageEvent.attempt_no > 1,
        )
        .group_by(RegistryObject.party_id)
    )
    for condition in _object_filter_conditions(filters):
        repeat_stmt = repeat_stmt.where(condition)
    repeat_rows = await session.execute(repeat_stmt)
    repeat_objects = {int(party_id): int(count) for party_id, count in repeat_rows.all() if party_id is not None}
    return counts, latest, repeat_objects


async def _party_object_flags(session: AsyncSession, parties: list[Party], filters: ReportFilters) -> dict[int, Counter[str]]:
    party_ids = [party.id for party in parties]
    if not party_ids:
        return {}
    no_object_by_party = {party.id: control_no_tokens(party.control_decree_without_object) for party in parties}
    no_decree_by_party = {party.id: control_no_tokens(party.control_object_without_decree) for party in parties}
    stmt = select(RegistryObject).where(RegistryObject.party_id.in_(party_ids), RegistryObject.status != "archived")
    for condition in _object_filter_conditions(filters):
        stmt = stmt.where(condition)
    result = await session.execute(stmt)
    counters: dict[int, Counter[str]] = defaultdict(Counter)
    for obj in result.scalars().all():
        if obj.party_id is None:
            continue
        party_id = int(obj.party_id)
        no_object = has_no_object_marker(obj.object_description) or any(control_token_matches_object(token, obj) for token in no_object_by_party.get(party_id, set()))
        no_decree = (not str(obj.decree_no or "").strip()) or any(control_token_matches_object(token, obj) for token in no_decree_by_party.get(party_id, set()))
        if no_object:
            counters[party_id]["no_object"] += 1
        if no_decree:
            counters[party_id]["no_decree"] += 1
        if _is_no_biomaterial(obj.object_description):
            counters[party_id]["no_biomaterial"] += 1
        if _is_burnt_bone(obj.object_description):
            counters[party_id]["burnt_bone"] += 1
    return counters


def _party_base_row(
    party: Party,
    object_count: int,
    stage_counts: dict[str, int],
    latest_stage: datetime | None,
    repeat_objects: int,
    flags: Counter[str],
) -> dict[str, Any]:
    progress = {}
    for stage in LAB_STAGES:
        done = int(stage_counts.get(stage, 0))
        progress[stage] = {"done": done, "total": object_count, "percent": _percent(done, object_count)}
    lab_percents = [progress[stage]["percent"] for stage in LAB_STAGES]
    readiness = round(sum(lab_percents) / len(lab_percents)) if lab_percents else 0
    lagging = min(LAB_STAGES, key=lambda stage: progress[stage]["percent"]) if object_count else None
    party_no_numeric = int(party.party_no) if str(party.party_no).isdigit() else None
    control_problem_count = _control_problem_count(party)
    return {
        "party_id": party.id,
        "party_no": party.party_no,
        "party_no_numeric": party_no_numeric,
        "party_no_sort": party_no_numeric if party_no_numeric is not None else -1,
        "case_year": party.case_year,
        "object_count": object_count,
        "stage_counts": stage_counts,
        "stage_progress": progress,
        "control_problem_count": control_problem_count,
        "control_status": _control_status(party),
        "readiness_percent": readiness,
        "lagging_stage": REPORT_STAGE_LABELS.get(lagging or "", lagging),
        "latest_change": (latest_stage or party.updated_at or party.created_at).isoformat() if (latest_stage or party.updated_at or party.created_at) else None,
        "status": "Архивная" if party.status == "archived" else "Активная",
        "repeat_stage_objects": repeat_objects,
        "no_object_count": int(flags.get("no_object", 0)),
        "no_decree_count": int(flags.get("no_decree", 0)),
        "no_biomaterial_count": int(flags.get("no_biomaterial", 0)),
        "burnt_bone_count": int(flags.get("burnt_bone", 0)),
    }


async def _party_rows(session: AsyncSession, filters: ReportFilters) -> list[dict[str, Any]]:
    parties = await _filtered_parties(session, filters)
    party_ids = [party.id for party in parties]
    object_counts = await _party_object_counts(session, party_ids, filters)
    stage_counts, latest_stage, repeat_objects = await _party_stage_counts(session, party_ids, filters)
    flags = await _party_object_flags(session, parties, filters)
    rows = [
        _party_base_row(
            party,
            object_counts.get(party.id, 0),
            stage_counts.get(party.id, {}),
            latest_stage.get(party.id),
            repeat_objects.get(party.id, 0),
            flags.get(party.id, Counter()),
        )
        for party in parties
    ]
    if not filters.include_empty_parties:
        rows = [row for row in rows if row["object_count"] > 0]
    if filters.only_problematic:
        rows = [row for row in rows if row["control_problem_count"] > 0 or row["readiness_percent"] < 100]
    return rows


async def get_reports_overview(session: AsyncSession, filters: ReportFilters) -> dict[str, Any]:
    rows = await _party_rows(session, filters)
    active_rows = [row for row in rows if row["status"] == "Активная"]
    stage_missing = {
        stage: sum(max(int(row["object_count"]) - int(row["stage_progress"][stage]["done"]), 0) for row in active_rows)
        for stage in LAB_STAGES
    }
    kpis = {
        "active_parties": len(active_rows),
        "total_objects": sum(int(row["object_count"]) for row in active_rows),
        "objects_in_work": sum(1 for row in active_rows if 0 < row["readiness_percent"] < 100),
        "problem_parties": sum(1 for row in active_rows if row["control_problem_count"] > 0 or row["readiness_percent"] < 100),
        "parties_without_control": sum(1 for row in active_rows if row["control_status"] == "Нет данных"),
        "objects_without_sample_prep": stage_missing["sample_prep"],
        "objects_without_milling": stage_missing["milling"],
        "objects_without_extraction": stage_missing["dna_extraction"],
        "objects_without_realtime": stage_missing["realtime"],
        "objects_without_pcr": stage_missing["pcr"],
        "objects_without_electrophoresis": stage_missing["electrophoresis"],
        "objects_without_analysis": stage_missing["analysis"],
        "objects_with_repeat_stages": sum(int(row["repeat_stage_objects"]) for row in active_rows),
        "objects_no_object": sum(int(row["no_object_count"]) for row in active_rows),
        "objects_no_decree": sum(int(row["no_decree_count"]) for row in active_rows),
        "objects_no_biomaterial": sum(int(row["no_biomaterial_count"]) for row in active_rows),
        "objects_burnt_bone": sum(int(row["burnt_bone_count"]) for row in active_rows),
    }
    sorted_rows = _sort_rows(rows, filters)
    return {"kpis": kpis, **_page_rows(sorted_rows, filters)}


async def get_party_control_report(session: AsyncSession, filters: ReportFilters, quick: str | None = None) -> dict[str, Any]:
    parties = await _filtered_parties(session, filters)
    party_ids = [party.id for party in parties]
    object_counts = await _party_object_counts(session, party_ids, filters)
    stage_counts, latest_stage, _repeat_objects = await _party_stage_counts(session, party_ids, filters)
    rows: list[dict[str, Any]] = []
    for party in parties:
        status = _control_status(party)
        row = {
            "party_id": party.id,
            "party_no": party.party_no,
            "party_no_numeric": int(party.party_no) if str(party.party_no).isdigit() else None,
            "party_no_sort": int(party.party_no) if str(party.party_no).isdigit() else -1,
            "case_year": party.case_year,
            "object_count": object_counts.get(party.id, 0),
            "control_actual_decrees": party.control_actual_decrees,
            "control_decree_without_object": party.control_decree_without_object,
            "control_object_without_decree": party.control_object_without_decree,
            "control_unidentified_rostov_no": party.control_unidentified_rostov_no,
            "control_need_recall": party.control_need_recall,
            "control_recalled": party.control_recalled,
            "problem_count": _control_problem_count(party),
            "control_status": status,
            "latest_change": (latest_stage.get(party.id) or party.updated_at or party.created_at).isoformat() if (latest_stage.get(party.id) or party.updated_at or party.created_at) else None,
            "status": "Архивная" if party.status == "archived" else "Активная",
            "stage_counts": stage_counts.get(party.id, {}),
        }
        rows.append(row)
    if not filters.include_empty_parties:
        rows = [row for row in rows if row["object_count"] > 0]
    if filters.only_problematic:
        rows = [row for row in rows if row["problem_count"] > 0 or row["control_status"] == "Нет данных"]
    quick_counts = {
        "": len(rows),
        "problem": sum(1 for row in rows if row["problem_count"] > 0),
        "critical": sum(1 for row in rows if row["control_status"] == "Критично"),
        "empty_control": sum(1 for row in rows if row["control_status"] == "Нет данных"),
        **{
            field: sum(1 for row in rows if _filled(row.get(field)))
            for field in CONTROL_FIELDS
        },
    }
    quick = quick or ""
    if quick == "problem":
        rows = [row for row in rows if row["problem_count"] > 0]
    elif quick == "critical":
        rows = [row for row in rows if row["control_status"] == "Критично"]
    elif quick in CONTROL_FIELDS:
        rows = [row for row in rows if _filled(row.get(quick))]
    elif quick == "empty_control":
        rows = [row for row in rows if row["control_status"] == "Нет данных"]
    sorted_rows = _sort_rows(rows, filters, default="problem_count")
    return {
        "items": _page_rows(sorted_rows, filters)["items"],
        "total": len(rows),
        "page": filters.page,
        "page_size": filters.page_size,
        "quick_counts": quick_counts,
    }


async def get_work_progress_report(session: AsyncSession, filters: ReportFilters, quick: str | None = None) -> dict[str, Any]:
    rows = await _party_rows(session, filters)
    quick_stage_map = {
        "no_sample_prep": "sample_prep",
        "no_milling": "milling",
        "no_extraction": "dna_extraction",
        "no_realtime": "realtime",
        "no_pcr": "pcr",
        "no_electrophoresis": "electrophoresis",
        "no_analysis": "analysis",
    }
    party_ids = [row["party_id"] for row in rows]
    pdf_parties = await _parties_with_pdf(session, party_ids, controls=False)
    control_pdf_parties = await _parties_with_pdf(session, party_ids, controls=True)
    quick_counts = {
        "": len(rows),
        **{
            quick_key: sum(
                1 for row in rows
                if row["stage_progress"][stage]["done"] < row["object_count"]
            )
            for quick_key, stage in quick_stage_map.items()
        },
        "repeat_analysis": sum(
            1 for row in rows
            if row["stage_counts"].get("analysis", 0) and row["repeat_stage_objects"] > 0
        ),
        "no_biomaterial": sum(1 for row in rows if row["no_biomaterial_count"] > 0),
        "burnt_bone": sum(1 for row in rows if row["burnt_bone_count"] > 0),
        "pdf": sum(1 for row in rows if row["party_id"] in pdf_parties),
        "control_pdf": sum(1 for row in rows if row["party_id"] in control_pdf_parties),
    }
    quick_stage = quick_stage_map.get(quick or "")
    if quick_stage:
        rows = [row for row in rows if row["stage_progress"][quick_stage]["done"] < row["object_count"]]
    elif quick == "repeat_analysis":
        rows = [row for row in rows if row["stage_counts"].get("analysis", 0) and row["repeat_stage_objects"] > 0]
    elif quick == "no_biomaterial":
        rows = [row for row in rows if row["no_biomaterial_count"] > 0]
    elif quick == "burnt_bone":
        rows = [row for row in rows if row["burnt_bone_count"] > 0]
    elif quick == "pdf":
        rows = [row for row in rows if row["party_id"] in pdf_parties]
    elif quick == "control_pdf":
        rows = [row for row in rows if row["party_id"] in control_pdf_parties]
    sorted_rows = _sort_rows(rows, filters, default="readiness_percent")
    return {**_page_rows(sorted_rows, filters), "quick_counts": quick_counts}


async def _parties_with_pdf(session: AsyncSession, party_ids: list[int], controls: bool) -> set[int]:
    if not party_ids:
        return set()
    if controls:
        result = await session.execute(
            select(ElectrophoresisControlFile.party_id).where(ElectrophoresisControlFile.party_id.in_(party_ids)).distinct()
        )
    else:
        result = await session.execute(
            select(RegistryObject.party_id)
            .join(ElectrophoresisResultFile, ElectrophoresisResultFile.object_id == RegistryObject.id)
            .where(RegistryObject.party_id.in_(party_ids), RegistryObject.status != "archived")
            .distinct()
        )
    return {int(party_id) for party_id in result.scalars().all() if party_id is not None}


def _period_key(period: str, dt: date | datetime | None, case_year: int | None = None) -> tuple[Any, str] | None:
    if not dt:
        return None
    if isinstance(dt, datetime):
        day = dt.date()
    else:
        day = dt
    bucket_year = case_year or day.year
    if period == "weekly":
        iso = day.isocalendar()
        start = day - timedelta(days=day.weekday())
        end = start + timedelta(days=6)
        return (bucket_year, iso.week), f"{bucket_year} · неделя {iso.week:02d}"
    if period == "monthly":
        return (bucket_year, day.month), f"{bucket_year}-{day.month:02d}"
    return (bucket_year,), str(bucket_year)


def _date_allowed(filters: ReportFilters, dt: date | datetime | None) -> bool:
    if not dt:
        return True
    day = dt.date() if isinstance(dt, datetime) else dt
    if filters.date_from and day < filters.date_from:
        return False
    if filters.date_to and day > filters.date_to:
        return False
    return True


async def get_period_statistics(session: AsyncSession, filters: ReportFilters, period: str) -> dict[str, Any]:
    party_stmt = select(Party.id, Party.created_at, Party.case_year).where(Party.status != "archived" if not filters.include_archived else True)
    object_stmt = select(RegistryObject.id, RegistryObject.party_id, RegistryObject.created_at, RegistryObject.case_year).where(RegistryObject.status != "archived")
    event_stmt = (
        select(StageEvent.stage_type, StageEvent.attempt_no, StageEvent.event_date, StageEvent.created_at, RegistryObject.party_id, RegistryObject.case_year)
        .join(RegistryObject, RegistryObject.id == StageEvent.object_id)
        .where(StageEvent.is_cancelled.is_(False), RegistryObject.status != "archived")
    )
    if filters.case_year:
        party_stmt = party_stmt.where(Party.case_year == filters.case_year)
        object_stmt = object_stmt.where(RegistryObject.case_year == filters.case_year)
        event_stmt = event_stmt.where(RegistryObject.case_year == filters.case_year)
    if filters.party_ids:
        party_stmt = party_stmt.where(Party.id.in_(filters.party_ids))
        object_stmt = object_stmt.where(RegistryObject.party_id.in_(filters.party_ids))
        event_stmt = event_stmt.where(RegistryObject.party_id.in_(filters.party_ids))
    party_rows = (await session.execute(party_stmt)).all()
    object_rows = (await session.execute(object_stmt)).all()
    event_rows = (await session.execute(event_stmt)).all()
    buckets: dict[tuple[Any, ...], dict[str, Any]] = {}

    def ensure(key: tuple[Any, ...], label: str) -> dict[str, Any]:
        row = buckets.setdefault(
            key,
            {
                "period_key": label,
                "year": key[0] if key else None,
                "week": key[1] if period == "weekly" and len(key) > 1 else None,
                "month": key[1] if period == "monthly" and len(key) > 1 else None,
                "new_parties": 0,
                "new_objects": 0,
                "stage_counts": {stage: 0 for stage in LAB_STAGES},
                "repeat_stage_events": 0,
                "control_problems": 0,
            },
        )
        return row

    for _id, created_at, _case_year in party_rows:
        if not _date_allowed(filters, created_at):
            continue
        keyed = _period_key(period, created_at, filters.case_year)
        if keyed:
            key, label = keyed
            ensure(key, label)["new_parties"] += 1
    for _id, _party_id, created_at, _case_year in object_rows:
        if not _date_allowed(filters, created_at):
            continue
        keyed = _period_key(period, created_at, filters.case_year)
        if keyed:
            key, label = keyed
            ensure(key, label)["new_objects"] += 1
    for stage_type, attempt_no, event_date, created_at, _party_id, _case_year in event_rows:
        event_dt = event_date or created_at
        if not _date_allowed(filters, event_dt):
            continue
        keyed = _period_key(period, event_dt, filters.case_year)
        if keyed:
            key, label = keyed
            row = ensure(key, label)
            if stage_type in LAB_STAGES:
                row["stage_counts"][stage_type] += 1
            if int(attempt_no or 1) > 1:
                row["repeat_stage_events"] += 1
    for party in await _filtered_parties(session, filters):
        keyed = _period_key(period, party.updated_at or party.created_at, filters.case_year)
        if keyed:
            key, label = keyed
            ensure(key, label)["control_problems"] += _control_problem_count(party)
    rows = sorted(buckets.values(), key=lambda row: row["period_key"], reverse=True)
    return _page_rows(rows, filters)


async def get_performer_statistics(session: AsyncSession, filters: ReportFilters) -> dict[str, Any]:
    stmt = (
        select(StageEventPerformer.raw_name, StageEventPerformer.role, StageEvent.stage_type, func.count(StageEvent.id))
        .join(StageEvent, StageEvent.id == StageEventPerformer.stage_event_id)
        .join(RegistryObject, RegistryObject.id == StageEvent.object_id)
        .where(StageEvent.is_cancelled.is_(False), RegistryObject.status != "archived")
        .group_by(StageEventPerformer.raw_name, StageEventPerformer.role, StageEvent.stage_type)
    )
    if filters.case_year:
        stmt = stmt.where(RegistryObject.case_year == filters.case_year)
    if filters.party_ids:
        stmt = stmt.where(RegistryObject.party_id.in_(filters.party_ids))
    if filters.employee_id:
        stmt = stmt.where(StageEventPerformer.employee_id == filters.employee_id)
    if filters.stage_type:
        stmt = stmt.where(StageEvent.stage_type == filters.stage_type)
    rows = await session.execute(stmt)
    by_person: dict[tuple[str, str], dict[str, Any]] = {}
    for raw_name, role, stage_type, count in rows.all():
        name = raw_name or "Не указан"
        key = (name, role or "")
        item = by_person.setdefault(
            key,
            {"employee": name, "role": role or "—", "stage_counts": {stage: 0 for stage in LAB_STAGES}, "total_actions": 0},
        )
        if stage_type in LAB_STAGES:
            item["stage_counts"][stage_type] += int(count)
        item["total_actions"] += int(count)
    result = sorted(by_person.values(), key=lambda row: row["total_actions"], reverse=True)
    return _page_rows(result, filters)


def build_report_workbook(title: str, columns: list[tuple[str, str]], rows: Iterable[dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    ws.append([label for _key, label in columns])
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in rows:
        ws.append([row.get(key) for key, _label in columns])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{max(ws.max_row, 1)}"
    for idx, (_key, label) in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = min(max(len(label) + 4, 14), 42)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    output = BytesIO()
    wb.save(output)
    return output.getvalue()
