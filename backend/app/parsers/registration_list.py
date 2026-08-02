from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


@dataclass
class RegistrationListColumn:
    column_index: int
    column_letter: str
    values: list[str]
    source_rows: list[int]


@dataclass
class RegistrationListPreview:
    sheet_name: str
    columns: list[RegistrationListColumn]
    warnings: list[str]


def _cell_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        text = str(int(value))
    else:
        text = str(value)
    text = " ".join(text.replace("\xa0", " ").split())
    return text or None


def parse_registration_list(path: str | Path) -> RegistrationListPreview:
    workbook = load_workbook(path, data_only=True, read_only=True)
    worksheet = workbook.worksheets[0]
    columns: list[RegistrationListColumn] = []
    warnings: list[str] = []

    for column_index in range(1, worksheet.max_column + 1):
        values: list[str] = []
        source_rows: list[int] = []
        for row_index in range(1, worksheet.max_row + 1):
            text = _cell_text(worksheet.cell(row=row_index, column=column_index).value)
            if not text:
                continue
            values.append(text)
            source_rows.append(row_index)
        if values:
            columns.append(
                RegistrationListColumn(
                    column_index=column_index,
                    column_letter=get_column_letter(column_index),
                    values=values,
                    source_rows=source_rows,
                )
            )

    if not columns:
        warnings.append("В файле не найдены заполненные колонки.")

    return RegistrationListPreview(sheet_name=worksheet.title, columns=columns, warnings=warnings)
