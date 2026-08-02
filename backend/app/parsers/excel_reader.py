from collections.abc import Iterable
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET
from zipfile import ZipFile


def _cell_to_python(value: Any) -> Any:
    if value is None:
        return ""
    return value


def read_workbook(path: str | Path) -> dict[str, list[list[Any]]]:
    """Read xls/xlsx/csv-like Excel files into sheet -> rows.

    python-calamine handles both old BIFF .xls and modern OOXML .xlsx.
    openpyxl is retained as a fallback for environments where calamine rejects
    a particular OOXML workbook.
    """

    path = Path(path)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        import csv

        with path.open("r", encoding="utf-8-sig", newline="") as fh:
            return {"CSV": [list(row) for row in csv.reader(fh)]}

    if suffix == ".xlsx":
        return _read_xlsx_physical_rows(path)

    try:
        from python_calamine import load_workbook

        workbook = load_workbook(str(path))
        sheets: dict[str, list[list[Any]]] = {}
        sheet_names = workbook.sheet_names() if callable(workbook.sheet_names) else workbook.sheet_names
        for sheet_name in sheet_names:
            sheet = workbook.get_sheet_by_name(sheet_name)
            rows = sheet.to_python()
            sheets[sheet_name] = [[_cell_to_python(cell) for cell in row] for row in rows]
        return sheets
    except Exception:
        if suffix != ".xlsx":
            raise

    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheets = {}
    for sheet in workbook.worksheets:
        sheets[sheet.title] = [
            [_cell_to_python(cell) for cell in row]
            for row in sheet.iter_rows(values_only=True)
        ]
    return sheets


def _read_xlsx_physical_rows(path: Path) -> dict[str, list[list[Any]]]:
    ns = {
        "m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        "pkg": "http://schemas.openxmlformats.org/package/2006/relationships",
    }

    def col_to_idx(ref: str) -> int:
        letters = "".join(ch for ch in ref if ch.isalpha())
        idx = 0
        for ch in letters:
            idx = idx * 26 + ord(ch.upper()) - 64
        return idx - 1

    def shared_strings(zf: ZipFile) -> list[str]:
        if "xl/sharedStrings.xml" not in zf.namelist():
            return []
        root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
        values: list[str] = []
        for si in root.findall("m:si", ns):
            values.append("".join((t.text or "") for t in si.iter(f"{{{ns['m']}}}t")))
        return values

    def cell_value(cell: ET.Element, shared: list[str]) -> Any:
        cell_type = cell.attrib.get("t")
        if cell_type == "s":
            value = cell.find("m:v", ns)
            if value is None or value.text is None:
                return ""
            return shared[int(value.text)]
        if cell_type == "inlineStr":
            return "".join(t.text or "" for t in cell.iter(f"{{{ns['m']}}}t"))
        value = cell.find("m:v", ns)
        return value.text if value is not None and value.text is not None else ""

    with ZipFile(path) as zf:
        shared = shared_strings(zf)
        workbook = ET.fromstring(zf.read("xl/workbook.xml"))
        rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
        rid_to_target = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels}
        result: dict[str, list[list[Any]]] = {}
        sheets_node = workbook.find("m:sheets", ns)
        if sheets_node is None:
            return result
        for sheet in sheets_node:
            name = sheet.attrib["name"]
            rid = sheet.attrib[f"{{{ns['r']}}}id"]
            target = rid_to_target[rid]
            sheet_path = f"xl/{target}" if not target.startswith("xl/") else target
            root = ET.fromstring(zf.read(sheet_path))
            rows: list[list[Any]] = []
            for row in root.findall(".//m:sheetData/m:row", ns):
                values: list[Any] = []
                for cell in row.findall("m:c", ns):
                    idx = col_to_idx(cell.attrib["r"])
                    while len(values) <= idx:
                        values.append("")
                    values[idx] = cell_value(cell, shared)
                rows.append(values)
            result[name] = rows
        return result


def row_value(row: list[Any], index: int) -> Any:
    if index < len(row):
        return row[index]
    return ""


def iter_nonempty_rows(rows: Iterable[list[Any]]) -> Iterable[tuple[int, list[Any]]]:
    for index, row in enumerate(rows, start=1):
        if any(str(cell).strip() for cell in row if cell is not None):
            yield index, row
