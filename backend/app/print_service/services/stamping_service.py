from __future__ import annotations

import io
import shutil
from collections import Counter
from pathlib import Path
from typing import Any

import fitz
from openpyxl import load_workbook


MM_TO_POINT = 72 / 25.4
MAX_LABEL_LENGTH = 100


class StampingValidationError(ValueError):
    pass


DEFAULT_STAMP_STYLE = {
    "corner": "top_left",
    "rotation": "none",
    "margin_x_mm": 10,
    "margin_y_mm": 8,
    "font_size": 12,
    "bold": False,
    "white_background": False,
    "border": False,
}


FONT_CANDIDATES = {
    False: [
        Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
        Path("/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf"),
    ],
    True: [
        Path("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
        Path("/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf"),
    ],
}


def default_stamp_config() -> dict[str, Any]:
    return {
        "enabled": False,
        "source": "manual",
        "text": "",
        "groups": {},
        "reject_duplicates": False,
        "allow_skip": False,
        "style": DEFAULT_STAMP_STYLE.copy(),
    }


def normalize_stamp_config(raw: dict[str, Any] | None) -> dict[str, Any]:
    config = default_stamp_config()
    if not raw:
        return config
    config["enabled"] = bool(raw.get("enabled"))
    config["source"] = str(raw.get("source") or "manual")[:30]
    config["text"] = str(raw.get("text") or "")
    config["groups"] = raw.get("groups") if isinstance(raw.get("groups"), dict) else {}
    config["reject_duplicates"] = bool(raw.get("reject_duplicates"))
    config["allow_skip"] = bool(raw.get("allow_skip"))
    style = raw.get("style") if isinstance(raw.get("style"), dict) else {}
    config["style"] = normalize_stamp_style(style)
    return config


def normalize_stamp_style(raw: dict[str, Any] | None) -> dict[str, Any]:
    raw = raw or {}
    style = DEFAULT_STAMP_STYLE.copy()
    if raw.get("corner") in {"top_left", "top_right", "bottom_left", "bottom_right"}:
        style["corner"] = raw["corner"]
    if raw.get("rotation") in {"none", "left_90"}:
        style["rotation"] = raw["rotation"]
    style["margin_x_mm"] = _bounded_float(raw.get("margin_x_mm"), 0, 80, style["margin_x_mm"])
    style["margin_y_mm"] = _bounded_float(raw.get("margin_y_mm"), 0, 80, style["margin_y_mm"])
    style["font_size"] = _bounded_float(raw.get("font_size"), 6, 48, style["font_size"])
    style["bold"] = bool(raw.get("bold"))
    style["white_background"] = bool(raw.get("white_background"))
    style["border"] = bool(raw.get("border"))
    return style


def _bounded_float(value: Any, minimum: float, maximum: float, default: float) -> float:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return default
    return max(minimum, min(maximum, number))


def parse_label_text(text: str) -> list[str]:
    labels: list[str] = []
    for raw in text.replace("\ufeff", "").splitlines():
        label = raw.strip()
        if not label:
            continue
        validate_label(label)
        labels.append(label)
    return labels


def parse_label_xlsx(path: Path, column: str | None = None) -> dict[str, Any]:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise StampingValidationError("Не удалось прочитать Excel-файл с метками") from exc
    sheet = workbook.active
    columns: list[dict[str, Any]] = []
    selected: list[str] = []
    selected_column = (column or "").strip().upper()
    for column_index in range(1, sheet.max_column + 1):
        letter = sheet.cell(row=1, column=column_index).column_letter
        values: list[str] = []
        for row_index in range(1, sheet.max_row + 1):
            value = sheet.cell(row=row_index, column=column_index).value
            if value is None:
                continue
            label = str(value).strip()
            if not label:
                continue
            validate_label(label)
            values.append(label)
        if values:
            columns.append(
                {
                    "column": letter,
                    "count": len(values),
                    "first": values[0],
                    "last": values[-1],
                    "labels": values if letter == selected_column else None,
                }
            )
            if letter == selected_column:
                selected = values
    workbook.close()
    return {"columns": columns, "selected_column": selected_column or None, "labels": selected}


def validate_label(label: str) -> None:
    if len(label) > MAX_LABEL_LENGTH:
        raise StampingValidationError("Метка длиннее 100 символов")
    if "\x00" in label:
        raise StampingValidationError("Метка содержит нулевой байт")
    for char in label:
        code = ord(char)
        if code < 32 and char not in {"\t"}:
            raise StampingValidationError("Метка содержит управляющий символ")


def apply_stamping_to_validation(validation: dict[str, Any], raw_config: dict[str, Any] | None) -> dict[str, Any]:
    config = normalize_stamp_config(raw_config)
    validation["stamping"] = {
        "config": config,
        "summary": {
            "documents": 0,
            "labels": 0,
            "applied": 0,
            "skipped": 0,
            "duplicate_count": 0,
            "ready": True,
        },
    }
    if not config["enabled"]:
        return validation
    if validation.get("mode") == "excel":
        return _apply_excel_stamping(validation, config)
    return _apply_text_stamping(validation, config)


def _apply_text_stamping(validation: dict[str, Any], config: dict[str, Any]) -> dict[str, Any]:
    entries = validation.get("entries") or []
    labels = parse_label_text(config.get("text") or "")
    errors, warnings, summary = validate_label_count(entries, labels, config)
    _attach_labels(entries, labels, config)
    validation["stamping"] = {"config": config, "summary": summary}
    validation.setdefault("warnings", []).extend(warnings)
    validation.setdefault("blocking_errors", []).extend(errors)
    validation["can_build"] = validation.get("can_build", False) and not errors
    return validation


def _apply_excel_stamping(validation: dict[str, Any], config: dict[str, Any]) -> dict[str, Any]:
    total_errors: list[str] = []
    total_warnings: list[str] = []
    stamped = 0
    skipped = 0
    duplicate_count = 0
    labels_total = 0
    groups_config = config.get("groups") or {}
    for group in validation.get("groups") or []:
        group_id = group.get("id") or group.get("column") or group.get("title")
        group_config = groups_config.get(group_id) or groups_config.get(group.get("column")) or {}
        if isinstance(group_config, str):
            group_config = {"text": group_config}
        labels = parse_label_text(str(group_config.get("text") or ""))
        errors, warnings, summary = validate_label_count(group["validation"].get("entries") or [], labels, config)
        _attach_labels(group["validation"].get("entries") or [], labels, config)
        group["stamping"] = {
            "labels_column": group_config.get("column") or "",
            "summary": summary,
        }
        total_errors.extend(f"{group['title']}: {error}" for error in errors)
        total_warnings.extend(f"{group['title']}: {warning}" for warning in warnings)
        stamped += summary["applied"]
        skipped += summary["skipped"]
        duplicate_count += summary["duplicate_count"]
        labels_total += summary["labels"]
    validation["stamping"] = {
        "config": config,
        "summary": {
            "documents": sum(group["validation"]["total"] for group in validation.get("groups") or []),
            "labels": labels_total,
            "applied": stamped,
            "skipped": skipped,
            "duplicate_count": duplicate_count,
            "ready": not total_errors,
        },
    }
    validation.setdefault("warnings", []).extend(total_warnings)
    validation.setdefault("blocking_errors", []).extend(total_errors)
    validation["can_build"] = validation.get("can_build", False) and not total_errors
    return validation


def validate_label_count(
    entries: list[dict[str, Any]], labels: list[str], config: dict[str, Any]
) -> tuple[list[str], list[str], dict[str, Any]]:
    errors: list[str] = []
    warnings: list[str] = []
    documents = len(entries)
    if len(labels) != documents:
        diff = documents - len(labels)
        if diff > 0:
            errors.append(
                f"Не совпадает количество документов и меток: документов {documents}, меток {len(labels)}. Не хватает {diff}."
            )
        else:
            errors.append(
                f"Не совпадает количество документов и меток: документов {documents}, меток {len(labels)}. Лишних меток: {abs(diff)}."
            )
    counts = Counter(label for label in labels if not is_skip_label(label, config))
    duplicate_count = sum(count - 1 for count in counts.values() if count > 1)
    if duplicate_count:
        message = f"Найдены повторяющиеся метки: {duplicate_count}"
        if config.get("reject_duplicates"):
            errors.append(message)
        else:
            warnings.append(message)
    skipped = sum(1 for label in labels if is_skip_label(label, config))
    summary = {
        "documents": documents,
        "labels": len(labels),
        "applied": max(0, len(labels) - skipped) if len(labels) == documents else 0,
        "skipped": skipped,
        "duplicate_count": duplicate_count,
        "ready": not errors,
    }
    return errors, warnings, summary


def _attach_labels(entries: list[dict[str, Any]], labels: list[str], config: dict[str, Any]) -> None:
    for index, entry in enumerate(entries):
        label = labels[index] if index < len(labels) else ""
        entry["stamp_label"] = label
        entry["stamp_skip"] = bool(label and is_skip_label(label, config))
        entry["stamp_applied"] = False


def is_skip_label(label: str, config: dict[str, Any]) -> bool:
    return bool(config.get("allow_skip")) and label.strip().upper() == "SKIP"


def stamp_pdf(input_pdf: Path, output_pdf: Path, label: str, style: dict[str, Any]) -> None:
    output_pdf.parent.mkdir(parents=True, exist_ok=True)
    document = fitz.open(input_pdf)
    try:
        for page in document:
            _stamp_page(page, label, style)
        document.save(output_pdf, garbage=4, deflate=True)
    finally:
        document.close()


def _stamp_page(page: fitz.Page, label: str, style: dict[str, Any]) -> None:
    rect = page.rect
    margin_x = style["margin_x_mm"] * MM_TO_POINT
    margin_y = style["margin_y_mm"] * MM_TO_POINT
    padding = 2 * MM_TO_POINT
    font_size = float(style["font_size"])
    font_file = _font_file(bool(style.get("bold"))) if _needs_external_font(label) else None
    font_name = "stampfont" if font_file else ("hebo" if style.get("bold") else "helv")
    text_width = _text_length(label, font_size, font_file, font_name)
    text_height = font_size * 1.35
    rotated_left = style.get("rotation") == "left_90"
    box_width = (text_height if rotated_left else text_width) + padding * 2
    box_height = (text_width if rotated_left else text_height) + padding * 2
    corner = style.get("corner") or "top_left"
    if "right" in corner:
        x0 = rect.x1 - margin_x - box_width
    else:
        x0 = rect.x0 + margin_x
    if "bottom" in corner:
        y0 = rect.y1 - margin_y - box_height
    else:
        y0 = rect.y0 + margin_y
    x0 = max(rect.x0, min(x0, rect.x1 - box_width))
    y0 = max(rect.y0, min(y0, rect.y1 - box_height))
    box = fitz.Rect(x0, y0, x0 + box_width, y0 + box_height)
    if style.get("white_background"):
        page.draw_rect(box, color=None, fill=(1, 1, 1), overlay=True)
    if style.get("border"):
        page.draw_rect(box, color=(0, 0, 0), width=0.5, overlay=True)
    if rotated_left:
        text_point = fitz.Point(box.x0 + padding + font_size, box.y1 - padding)
        page.insert_text(
            text_point,
            label,
            fontsize=font_size,
            fontname=font_name,
            fontfile=str(font_file) if font_file else None,
            color=(0, 0, 0),
            rotate=90,
            overlay=True,
        )
        return
    text_point = fitz.Point(box.x0 + padding, box.y0 + padding + font_size)
    page.insert_text(
        text_point,
        label,
        fontsize=font_size,
        fontname=font_name,
        fontfile=str(font_file) if font_file else None,
        color=(0, 0, 0),
        overlay=True,
    )


def _font_file(bold: bool) -> Path | None:
    for path in FONT_CANDIDATES[bold]:
        if path.exists():
            return path
    return None


def _needs_external_font(label: str) -> bool:
    try:
        label.encode("latin-1")
    except UnicodeEncodeError:
        return True
    return False


def _text_length(label: str, font_size: float, font_file: Path | None, font_name: str) -> float:
    if font_file:
        try:
            return fitz.Font(fontfile=str(font_file)).text_length(label, fontsize=font_size)
        except Exception:
            pass
    try:
        return fitz.get_text_length(label, fontname=font_name, fontsize=font_size)
    except Exception:
        return max(len(label), 1) * font_size * 0.6


def stamp_entries(
    entries: list[dict[str, Any]],
    pdf_paths: list[Path],
    job_dir: Path,
    stamp_config: dict[str, Any] | None,
) -> tuple[list[Path], dict[str, Any]]:
    config = normalize_stamp_config(stamp_config)
    if not config.get("enabled"):
        return pdf_paths, {"enabled": False, "applied": 0, "skipped": 0}
    output_paths: list[Path] = []
    applied = 0
    skipped = 0
    style = config["style"]
    for entry, pdf_path in zip(entries, pdf_paths, strict=True):
        label = entry.get("stamp_label") or ""
        if not label or entry.get("stamp_skip"):
            if entry.get("stamp_skip"):
                skipped += 1
                entry["stamp_applied"] = False
            output_paths.append(pdf_path)
            continue
        output_pdf = job_dir / "stamped" / f"{entry['order']:04d}.pdf"
        stamp_pdf(pdf_path, output_pdf, label, style)
        entry["stamp_applied"] = True
        entry["stamped_pdf_path"] = str(output_pdf.relative_to(job_dir))
        applied += 1
        output_paths.append(output_pdf)
    return output_paths, {"enabled": True, "applied": applied, "skipped": skipped}


def copy_preview_with_stamp(source_pdf: Path, output_pdf: Path, label: str, style: dict[str, Any]) -> None:
    if label:
        stamp_pdf(source_pdf, output_pdf, label, style)
    else:
        output_pdf.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(source_pdf, output_pdf)


def render_preview_png(source_pdf: Path, output_png: Path) -> None:
    output_png.parent.mkdir(parents=True, exist_ok=True)
    document = fitz.open(source_pdf)
    try:
        page = document[0]
        pixmap = page.get_pixmap(matrix=fitz.Matrix(1.2, 1.2), alpha=False)
        pixmap.save(output_png)
    finally:
        document.close()


def labels_to_text(labels: list[str]) -> str:
    return "\n".join(labels)


def text_to_upload_file_bytes(text: str) -> io.BytesIO:
    return io.BytesIO(text.encode("utf-8"))
