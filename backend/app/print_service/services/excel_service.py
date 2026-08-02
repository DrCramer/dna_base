from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app.print_service.services.matching_service import contains_number_token, match_documents, normalize_text


class ExcelValidationError(ValueError):
    pass


def _document_contains(documents: list[dict[str, Any]], number: str) -> bool:
    normalized = normalize_text(number)
    return any(contains_number_token(normalize_text(Path(doc["original_name"]).stem), normalized) for doc in documents)


def parse_excel_sequences(path: Path, documents: list[dict[str, Any]]) -> list[dict[str, Any]]:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise ExcelValidationError("Не удалось прочитать Excel-файл") from exc

    sheet = workbook.active
    groups: list[dict[str, Any]] = []
    for column_index in range(1, sheet.max_column + 1):
        raw_values: list[str] = []
        for row_index in range(1, sheet.max_row + 1):
            value = sheet.cell(row=row_index, column=column_index).value
            if value is None:
                continue
            text = str(value).strip()
            if text:
                raw_values.append(text)
        if not raw_values:
            continue

        title = f"Столбец {get_column_letter(column_index)}"
        numbers = raw_values
        if len(raw_values) > 1 and not _document_contains(documents, raw_values[0]):
            later_matches = any(_document_contains(documents, value) for value in raw_values[1:])
            if later_matches:
                title = raw_values[0]
                numbers = raw_values[1:]

        if not numbers:
            continue
        groups.append(
            {
                "id": f"col_{get_column_letter(column_index)}",
                "title": title,
                "column": get_column_letter(column_index),
                "sequence": "\n".join(numbers),
                "count": len(numbers),
            }
        )
    workbook.close()

    if not groups:
        raise ExcelValidationError("В Excel-файле не найдено ни одного столбца с номерами")
    return groups


def match_excel_groups(path: Path, documents: list[dict[str, Any]]) -> dict[str, Any]:
    groups = parse_excel_sequences(path, documents)
    used_doc_ids: set[str] = set()
    blocking_errors: list[str] = []
    warnings: list[str] = []
    validated_groups: list[dict[str, Any]] = []

    for group in groups:
        validation = match_documents(group["sequence"], documents)
        validation["unused_documents"] = []
        for entry in validation["entries"]:
            if entry.get("doc_id"):
                used_doc_ids.add(entry["doc_id"])
        if validation["blocking_errors"]:
            blocking_errors.extend(
                f"{group['title']}: {error}" for error in validation["blocking_errors"]
            )
        warnings.extend(f"{group['title']}: {warning}" for warning in validation["warnings"])
        validated_groups.append({**group, "validation": validation})

    unused = [doc for doc in documents if doc["id"] not in used_doc_ids]
    if unused:
        warnings.append(f"Не используются в итоговых PDF: {len(unused)}")

    return {
        "mode": "excel",
        "groups": validated_groups,
        "unused_documents": unused,
        "warnings": warnings,
        "blocking_errors": blocking_errors,
        "can_build": len(blocking_errors) == 0,
        "total_groups": len(validated_groups),
        "total": sum(group["validation"]["total"] for group in validated_groups),
    }
