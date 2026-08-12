from datetime import date
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

from app.parsers.normalization import parse_date
from app.parsers.normalization import extract_party_no
from app.parsers.registration_list import parse_registration_list
from app.parsers.registry import REGISTRY_HEADERS, _extract_stage_events, _header_indexes, parse_registry
from app.services.export import build_registry_workbook
from app.services.stages import _merge_sample_prep_specs, registry_event_specs


REGISTRY_FILE = "/home/drcramer/.codex/attachments/9b019c0b-c894-492d-bb73-932663ab194d/№ 148 реестр.xlsx"
REGISTRY_132_FILE = "/home/drcramer/.codex/attachments/08902fc4-8716-4c7c-9cf4-01ecaab9bf22/№ 132 реестр.xlsx"
REGISTRY_109_FILE = "/home/drcramer/.codex/attachments/d2de6e2b-916e-4b35-9cc5-69b865c645f2/№ 109 Реестр.xlsx"
REGISTRY_108_STORED_FILE = "/app/data/registry/7548c292688b469cbf79226494645541.xlsx"
REGISTRY_111_STORED_FILE = "/app/data/registry/cdaeb5c2d0554eeeb0de2a8f48e77ed6.xlsx"
REGISTRY_114_STORED_FILE = "/app/data/registry/06c6c385832c4c7889697f2747f675ec.xlsx"
REGISTRY_160_STORED_FILE = "/app/data/registry/0ea98152450846c38f528615842ead17.xlsx"
REGISTRATION_LIST_FILE = "/home/drcramer/.codex/attachments/abeb7199-cb0b-414f-be34-35f57f17c81b/общий список.xlsx"


def test_registry_import_detects_real_100_rows_and_skips_noise():
    preview = parse_registry(REGISTRY_FILE)

    assert preview.sheet_name == "список объектов"
    assert len(preview.headers) == 65
    assert len(preview.rows) == 100
    assert preview.rows[0]["decree_no"] == "3770-2026"
    assert preview.rows[0]["rcsme_reg_no"] == "3770-1"
    assert preview.rows[-1]["decree_no"] == "3869-2026"
    assert preview.rows[-1]["rcsme_reg_no"] == "3869-1"
    assert all(row["source_row_number"] != 1046899 for row in preview.rows)
    assert all(row["source_row_number"] not in {105, 108, 111, 114, 117, 120} for row in preview.rows)


def test_registry_number_bases_are_searchable():
    preview = parse_registry(REGISTRY_FILE)
    first = preview.rows[0]

    assert first["decree_no_base"] == "3770"
    assert first["rcsme_reg_no_base"] == "3770"


def test_registry_132_extracts_stage_events_from_repeated_blocks():
    preview = parse_registry(REGISTRY_132_FILE)
    first = preview.rows[0]
    events = first["stage_events"]

    assert first["rcsme_reg_no"] == "2205-2026"
    assert any(event["block"] == "prep_photo" for event in events)
    assert any(event["block"] == "dna_extraction_1" for event in events)
    assert not any(event["block"] == "dna_extraction_2" for event in events)
    assert not any(event["table"] == "pcr_events" for event in events)
    assert not any(event["table"] == "electrophoresis_events" for event in events)

    analysis_events = [event for event in events if event["table"] == "electrophoresis_analysis_events"]
    assert [event["data"]["attempt_no"] for event in analysis_events] == [1, 2]
    assert analysis_events[0]["data"]["analysis_date"] == date(2026, 6, 10)
    assert analysis_events[1]["data"]["analysis_date"] == date(2026, 6, 23)
    assert all(event["raw_json"]["block"] == event["block"] for event in events)


def test_registry_132_excel_serial_analysis_date():
    preview = parse_registry(REGISTRY_132_FILE)
    second = preview.rows[1]
    analysis = [event for event in second["stage_events"] if event["block"] == "electrophoresis_analysis_1"][0]

    assert second["rcsme_reg_no"] == "2206-2026"
    assert analysis["data"]["analysis_date"] == date(2026, 6, 10)


def test_registry_109_extracts_late_pcr_electrophoresis_and_analysis():
    preview = parse_registry(REGISTRY_109_FILE)
    first = preview.rows[0]
    events = first["stage_events"]

    pcr = [event for event in events if event["table"] == "pcr_events"][0]
    electrophoresis = [event for event in events if event["table"] == "electrophoresis_events"][0]
    analysis = [event for event in events if event["block"] == "electrophoresis_analysis_1"][0]

    assert len(preview.headers) == 74
    assert len(preview.rows) == 101
    assert first["rcsme_reg_no"] == "101-1"
    assert pcr["data"]["pcr_date"] == date(2026, 2, 2)
    assert pcr["data"]["locus_panel"] == "GlobalFiltr"
    assert pcr["data"]["normalization_performer"] == "Попова М.М."
    assert pcr["data"]["pcr_performer"] == "Попова М.М."
    assert electrophoresis["data"]["electrophoresis_date"] == date(2026, 2, 3)
    assert electrophoresis["data"]["pipetting_method"] == "Ручной"
    assert electrophoresis["data"]["performer_1"] == "Попова М.М."
    assert analysis["data"]["analysis_date"] == date(2026, 2, 4)
    assert analysis["data"]["performer"] == "Смоляницкая А.И."

    specs = [spec for event in events for spec in registry_event_specs(event)]
    pcr_spec = [spec for spec in specs if spec["stage_type"] == "pcr"][0]
    electrophoresis_spec = [spec for spec in specs if spec["stage_type"] == "electrophoresis"][0]
    analysis_spec = [spec for spec in specs if spec["stage_type"] == "analysis"][0]

    assert pcr_spec["detail_data"]["normalization_performers"] == ["Попова М.М."]
    assert pcr_spec["detail_data"]["pcr_performers"] == ["Попова М.М."]
    assert pcr_spec["performers"] == [
        {"raw_name": "Попова М.М.", "role": "normalization"},
        {"raw_name": "Попова М.М.", "role": "pcr"},
    ]
    assert electrophoresis_spec["detail_data"]["performers"] == ["Попова М.М."]
    assert electrophoresis_spec["performers"][0] == {"raw_name": "Попова М.М.", "role": "performer"}
    assert analysis_spec["detail_data"]["analysis_date"] == date(2026, 2, 4)
    assert analysis_spec["performers"] == [{"raw_name": "Смоляницкая А.И.", "role": "analysis"}]


@pytest.mark.parametrize(
    ("path", "expected_first", "expected_analysis_date", "expected_performer", "expected_headers"),
    [
        (REGISTRY_108_STORED_FILE, "10-1", date(2026, 1, 29), "Непомнящая Л.В.", 74),
        (REGISTRY_111_STORED_FILE, "301-1", date(2026, 2, 9), "Смоляницкий А.Г.", 66),
    ],
)
def test_registry_stored_formats_extract_real_analysis_pairs(
    path: str,
    expected_first: str,
    expected_analysis_date: date,
    expected_performer: str,
    expected_headers: int,
):
    if not Path(path).exists():
        pytest.skip(f"stored registry fixture is not available: {path}")

    preview = parse_registry(path)
    first = preview.rows[0]
    analysis_events = [event for event in first["stage_events"] if event["table"] == "electrophoresis_analysis_events"]

    assert len(preview.headers) == expected_headers
    assert first["rcsme_reg_no"] == expected_first
    assert analysis_events[0]["data"]["analysis_date"] == expected_analysis_date
    assert analysis_events[0]["data"]["performer"] == expected_performer
    assert all(event["data"]["performer"] != "Ручной" for event in analysis_events)
    assert all(parse_date(event["data"]["performer"]) is None for event in analysis_events if event["data"]["performer"])


def test_excel_serial_date_46196():
    assert parse_date(46196) == date(2026, 6, 23)


def test_dash_separated_registry_date():
    assert parse_date("21-05-2026") == date(2026, 5, 21)
    assert parse_date("21-05-26") == date(2026, 5, 21)


def test_registry_114_dash_dates_restore_first_analysis():
    if not Path(REGISTRY_114_STORED_FILE).exists():
        pytest.skip(f"stored registry fixture is not available: {REGISTRY_114_STORED_FILE}")

    preview = parse_registry(REGISTRY_114_STORED_FILE)
    by_no = {row["rcsme_reg_no"]: row for row in preview.rows}

    first = by_no["511-1"]
    first_pcr = [
        event for event in first["stage_events"] if event["table"] == "pcr_events"
    ][0]
    first_electrophoresis = [
        event for event in first["stage_events"] if event["table"] == "electrophoresis_events"
    ][0]
    first_analysis = [
        event for event in first["stage_events"] if event["table"] == "electrophoresis_analysis_events"
    ][0]

    repeated = by_no["519-1"]
    repeated_analysis = [
        event for event in repeated["stage_events"] if event["table"] == "electrophoresis_analysis_events"
    ]

    assert first_pcr["data"]["pcr_date"] == date(2026, 5, 19)
    assert first_pcr["data"]["pcr_performer"] == "Непомнящая Л.В."
    assert first_electrophoresis["data"]["electrophoresis_date"] == date(2026, 5, 20)
    assert first_electrophoresis["data"]["sequencer"] == "3500"
    assert first_electrophoresis["data"]["performer_1"] == "Непомнящая Л.В."
    assert first_analysis["data"]["analysis_date"] == date(2026, 5, 21)
    assert first_analysis["data"]["performer"] == "Непомнящая Л.В."
    assert [(event["data"]["attempt_no"], event["data"]["analysis_date"]) for event in repeated_analysis] == [
        (1, date(2026, 5, 21)),
        (2, date(2026, 6, 4)),
    ]


def test_registry_160_preserves_registry_filled_by_in_sample_prep_specs():
    if not Path(REGISTRY_160_STORED_FILE).exists():
        pytest.skip(f"stored registry fixture is not available: {REGISTRY_160_STORED_FILE}")

    preview = parse_registry(REGISTRY_160_STORED_FILE)
    first = preview.rows[0]
    sample_prep_specs = [
        spec
        for event in first["stage_events"]
        for spec in registry_event_specs(event)
        if spec["stage_type"] == "sample_prep"
    ]

    assert first["rcsme_reg_no"] == "4842-1"
    assert first["registry_filled_by"] == "Нечитайлова Н.А."
    assert sample_prep_specs
    merged = _merge_sample_prep_specs(sample_prep_specs, first["registry_filled_by"])
    assert merged
    assert merged["detail_data"]["registry_filled_by"] == "Нечитайлова Н.А."


def test_sample_prep_merge_uses_row_level_registry_filled_by():
    merged = _merge_sample_prep_specs(
        [
            {
                "stage_type": "sample_prep",
                "event_date": date(2026, 7, 6),
                "comment": None,
                "detail_data": {"photo_performers": ["Нечитайлова Н.А."]},
                "performers": [],
            },
            {
                "stage_type": "sample_prep",
                "event_date": date(2026, 7, 7),
                "comment": None,
                "detail_data": {"washing_performers": ["Давыдов В.Е."]},
                "performers": [],
            },
        ],
        "Нечитайлова Н.А.",
    )

    assert merged
    assert merged["detail_data"]["registry_filled_by"] == "Нечитайлова Н.А."
    assert merged["detail_data"]["photo_performers"] == ["Нечитайлова Н.А."]
    assert merged["detail_data"]["washing_performers"] == ["Давыдов В.Е."]


def test_sample_prep_merge_creates_event_from_registry_filled_by_only():
    merged = _merge_sample_prep_specs([], "Нечитайлова Н.А.")

    assert merged
    assert merged["stage_type"] == "sample_prep"
    assert merged["detail_data"]["registry_filled_by"] == "Нечитайлова Н.А."
    assert merged["detail_data"]["photo_performers"] == []


def test_registry_parser_accepts_short_registry_filled_by_header():
    indexes = _header_indexes(["№ постановления", "№ рег РЦСМЭ", "Заполнение реестра"])

    assert indexes["registry_filled_by"] == 2


def test_registry_analysis_preserves_genotype_and_performer_without_date():
    headers = list(REGISTRY_HEADERS)
    row = [None] * len(headers)
    electrophoresis_start = headers.index("Дата электрофореза")
    analysis_start = headers.index("Дата анализа фореза")
    row[electrophoresis_start + 5] = "GT-01"
    row[analysis_start + 1] = "Иванов И.И."

    events = _extract_stage_events(headers, row)
    analysis = next(event for event in events if event["table"] == "electrophoresis_analysis_events")
    spec = registry_event_specs(analysis)[0]

    assert analysis["data"]["analysis_date"] is None
    assert analysis["data"]["performer"] == "Иванов И.И."
    assert analysis["data"]["genotype"] == "GT-01"
    assert spec["detail_data"]["genotype"] == "GT-01"
    assert spec["performers"] == [{"raw_name": "Иванов И.И.", "role": "analysis"}]


def test_registry_keeps_extraction_and_realtime_comments_in_their_own_stages():
    headers = list(REGISTRY_HEADERS)
    row = [None] * len(headers)
    row[34] = "Комментарий RealTime 1"
    row[38] = "Комментарий выделения 2"
    row[43] = "Комментарий RealTime 2"

    events = _extract_stage_events(headers, row)
    extraction_1 = next(event for event in events if event["block"] == "dna_extraction_1")
    extraction_2 = next(event for event in events if event["block"] == "dna_extraction_2")
    specs_1 = registry_event_specs(extraction_1)
    specs_2 = registry_event_specs(extraction_2)

    assert [spec["stage_type"] for spec in specs_1] == ["realtime"]
    assert specs_1[0]["comment"] == "Комментарий RealTime 1"
    assert [spec["stage_type"] for spec in specs_2] == ["dna_extraction", "realtime"]
    assert specs_2[0]["comment"] == "Комментарий выделения 2"
    assert specs_2[1]["comment"] == "Комментарий RealTime 2"


def test_registration_list_parser_reads_party_columns():
    if not Path(REGISTRATION_LIST_FILE).exists():
        pytest.skip(f"registration list fixture is not available: {REGISTRATION_LIST_FILE}")

    preview = parse_registration_list(REGISTRATION_LIST_FILE)

    assert preview.sheet_name == "Лист1"
    assert len(preview.columns) == 13
    assert sum(len(column.values) for column in preview.columns) == 1264
    assert [column.column_letter for column in preview.columns] == ["A", "C", "E", "G", "I", "K", "M", "O", "Q", "S", "U", "W", "Y"]
    assert [len(column.values) for column in preview.columns[:-1]] == [100] * 12
    assert len(preview.columns[-1].values) == 64


def test_extract_party_no_from_filename():
    assert extract_party_no("143 партия.xlsx") == "143"
    assert extract_party_no("партия 143 реестр.xlsx") == "143"
    assert extract_party_no("№ 148 реестр.xlsx") == "148"


def test_export_contains_65_registry_columns(tmp_path):
    obj = SimpleNamespace(
        registry_row_no="1",
        party_no="148",
        intake_date=None,
        decision_date=None,
        investigator=None,
        incoming_no=None,
        decree_no="3770-2026",
        object_description="",
        external_military_no="сс4169",
        extraction_note=None,
        box_no="2",
        packages_count=None,
        rcsme_reg_no="3770-1",
        object_type=None,
        extracted_before=None,
        not_extracted_before=None,
        registry_filled_by=None,
        raw_registry_json={},
    )
    content = build_registry_workbook([obj])
    path = tmp_path / "export.xlsx"
    path.write_bytes(content)
    wb = load_workbook(path)
    ws = wb["список объектов"]

    assert ws.max_column == 65
    assert [ws.cell(1, col).value for col in range(1, 66)] == REGISTRY_HEADERS
    assert ws.cell(2, 7).value == "3770-2026"
    assert ws.cell(2, 13).value == "3770-1"
